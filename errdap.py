from selenium import webdriver
import time
import openpyxl
import pyautogui as pag

from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.remote.webdriver import WebDriver as RemoteWebDriver
chromeOptions = webdriver.ChromeOptions()
chromeOptions.add_experimental_option("excludeSwitches", ["enable-logging"])
driverService = Service('D:/GRIIDC_Automation-main/GRIIDC_Automation-main/chromedriver.exe') 
driver = webdriver.Chrome(service=driverService,options=chromeOptions)
filePath = "D:/ERRDAP/dataset.xlsx"
workBook = openpyxl.load_workbook(filePath)
workbookSheet = workBook.active
sheetRow = workbookSheet.max_row
showColoumn = workbookSheet.max_column
x = []
y=0
for i in range(53, 132):
    for c in range(2, 3):
        y=y+1
        cellObj = workbookSheet.cell(row = i, column =c)
        st = str(cellObj.value)
        driver.get('https://erddap.gcoos.org/erddap/tabledap/'+st+'.html')
        time.sleep(5)
        pag.hotkey('ctrl', 'f')
        pag.write('source')
        time.sleep(3)
        pag.hotkey('ctrl', 'f')
        pag.write('cdm_data_type')
        # pag.hotkey('ctrl', 'shft','r')
        time.sleep(5)
driver.close()
print(y)
        # a = ActionChains(driver)
        # # perform the ctrl+f pressing action
        # a.key_down(Keys.CONTROL).send_keys('F').key_up(Keys.CONTROL).perform()